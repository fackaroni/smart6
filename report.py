import pandas as pd
import datetime
import pytz
from openpyxl.styles import PatternFill, Border, Side, Font

def calculate_results(data):
    data['solids_oven_average'] = (data['solids_oven1'] + data['solids_oven2']) / 2
    data['deviation'] = abs(data['solids_oven_average'] - data['solids_smart'])

    grouped = data.groupby(['material_number', 'method'])[['deviation', 'drying_time', 'ew']].agg({
        'deviation': ['count', 'mean', 'std'],
        'drying_time': 'mean',
        'ew': 'mean'
    })

    min_deviation_methods = grouped[('deviation', 'mean')].groupby(level=0).idxmin().to_dict()

    cache = {}
    results_list = []

    for (material_number, method), row in grouped.iterrows():
        key = (material_number, method)
        if key not in cache:
            cache[key] = data[(data['material_number'] == material_number) & (data['method'] == method)]
        material = cache[key]['material'].mode().iloc[0]
        passed = 'x' if row[('deviation', 'count')] >= 5 and row[('deviation', 'mean')] <= 1 else ''
        deviation_values = get_deviation_values(cache[key])
        ew_values = get_ew_values(cache[key])
        
        lowest_deviation = 'x' if min_deviation_methods[material_number] == key else ''

        results_list.append({
            'material_number': material_number,
            'material': material,
            'method': method,
            'lowest_deviation': lowest_deviation,
            'count': row[('deviation', 'count')],
            'avg_deviation': row[('deviation', 'mean')],
            'std_deviation': row[('deviation', 'std')],
            'avg_drying_time': row[('drying_time', 'mean')],
            'avg_ew': row[('ew', 'mean')],
            'deviation_values': deviation_values,
            'ew_values': ew_values
        })

    calculation_results = pd.DataFrame(results_list)
    return calculation_results

def get_deviation_values(filtered_data):
    deviation_values = filtered_data['deviation']
    return ', '.join(map("{:.2f}".format, deviation_values))

def get_ew_values(filtered_data):
    ew_values = filtered_data['ew']
    return ', '.join(map("{:.2f}".format, ew_values))

def get_current_time_in_gmt_plus_2():
    tz = pytz.timezone('Etc/GMT-2')
    return datetime.datetime.now(tz)

def adjust_column_widths(worksheet):
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

def save_results_to_excel(filtered_results, sheet_name='Results'):
    date_string = get_current_time_in_gmt_plus_2().strftime("%Y%m%d-%H%M")
    file_path = f'smart-report-{date_string}.xlsx'
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        filtered_results.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        passing_material_numbers = filtered_results[(filtered_results['count'] >= 5) & (filtered_results['avg_deviation'] <= 1)]['material_number'].unique()
        style_cells_in_excel(worksheet, filtered_results)
        adjust_column_widths(worksheet)
    return file_path

def style_cells_in_excel(worksheet, data):
    grey_border = Border(left=Side(style='thin', color='D3D3D3'),
                         right=Side(style='thin', color='D3D3D3'),
                         top=Side(style='thin', color='D3D3D3'),
                         bottom=Side(style='thin', color='D3D3D3'))

    for idx, row in data.iterrows():
        for col_num, column_name in enumerate(data.columns):
            cell = worksheet.cell(row=idx+2, column=col_num+1)
            styles = get_cell_styles(row, column_name)
            
            cell.font = Font(color=styles["font_color"], bold=styles["bold"])
            cell.fill = PatternFill(start_color=styles["bg_color"], end_color=styles["bg_color"], fill_type="solid")
            cell.border = grey_border

def get_cell_styles(row, column_name):
    count, avg_dev, avg_ew, avg_drying_time = row['count'], row['avg_deviation'], row['avg_ew'], row['avg_drying_time']

    styles = {
        "font_color": '000000',
        "bg_color": 'FFFFFF',
        "bold": False
    }

    if (count < 3 and avg_dev < 1): 
        pass
    elif count >= 3 and avg_dev < 1: 
        styles["bg_color"] = "98FB98"
    elif avg_dev > 1 and avg_ew > 1.7 and count < 3  or (avg_dev < 2 and avg_ew == 0):
        styles["bg_color"] = "FFA07A"
    elif (avg_dev > 1 and avg_ew < 1.7) or (avg_dev > 1 and count > 3) or (avg_dev > 2 and avg_ew == 0):
        styles["font_color"] = 'FF0000'

    if column_name == 'avg_drying_time' and avg_drying_time >= 7:
        styles["bold"] = True
    elif column_name == 'avg_ew' and avg_ew >= 1.7:
        styles["bold"] = True

    if avg_drying_time >= 7 and avg_ew >= 1.7:
        if column_name in ['avg_drying_time', 'avg_ew']:
            styles["font_color"] = '0000FF' 

    return styles

# Usage
data = pd.read_excel('smart6.xlsx', engine='openpyxl')

#----Lines to filter output for report based on conditions---

#data = data[(data['ew'] < 1.5) & (data['ew'] != 0)]


results = calculate_results(data)
save_results_to_excel(results)
